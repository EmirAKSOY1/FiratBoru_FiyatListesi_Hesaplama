from openpyxl import Workbook,load_workbook
from openpyxl.styles import PatternFill

kdv=0.18
iskonto=0.5

eski = load_workbook("fiyatlar.xlsx")
eski_ws = eski.active

yeni= Workbook()

yeni_ws = yeni.active
yeni_ws.title = "İskonto ve KDV uygulanmış hali"
yeni_ws = yeni.create_sheet("yeni")


yeni_ws["E4"]="Fiyat"
yeni_ws["D4"]="Ürün Adı"
yeni_ws.sheet_properties.tabColor = "1072BA"

fill_cell1 = PatternFill(patternType='solid', fgColor='E0E0E0')
fill_cell2 = PatternFill(patternType='solid', fgColor='808080')

for i in range(5,200):
    isim_hucre="D"+str(i)
    fiyat_hucre="F"+str(i)
    yeni_fiyat_hucre="E"+str(i)
    isim=eski_ws[isim_hucre].value
    fiyat=float(eski_ws[fiyat_hucre].value)
    fiyat=(fiyat*kdv+fiyat)-(fiyat*iskonto)
    print(isim,"----->",fiyat)

    yeni_ws[isim_hucre] = isim
    yeni_ws[yeni_fiyat_hucre] = fiyat
    if i%2==0:
        yeni_ws[yeni_fiyat_hucre].fill=fill_cell1

    else:
        yeni_ws[yeni_fiyat_hucre].fill=fill_cell2
yeni.save("yeni_fiyatlar.xlsx")


